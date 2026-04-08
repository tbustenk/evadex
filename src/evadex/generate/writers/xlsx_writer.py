"""XLSX writer for evadex generate.

Structure:
  - Sheet 0: Summary (totals, evasion stats, category breakdown)
  - One sheet per category (columns: #, Embedded Text, Plain Value, Variant Value,
    Technique, Generator, Has Keywords)
"""
from __future__ import annotations

import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from evadex.generate.generator import GeneratedEntry
from evadex.core.result import PayloadCategory


_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
_EVASION_FILL = PatternFill("solid", fgColor="FFF2CC")

_SECTION_TITLES: dict[PayloadCategory, str] = {
    PayloadCategory.CREDIT_CARD: "Credit Cards",
    PayloadCategory.SSN:         "SSN",
    PayloadCategory.SIN:         "SIN",
    PayloadCategory.IBAN:        "IBAN",
    PayloadCategory.SWIFT_BIC:   "SWIFT-BIC",
    PayloadCategory.ABA_ROUTING: "ABA Routing",
    PayloadCategory.BITCOIN:     "Bitcoin",
    PayloadCategory.ETHEREUM:    "Ethereum",
    PayloadCategory.US_PASSPORT: "US Passport",
    PayloadCategory.AU_TFN:      "AU TFN",
    PayloadCategory.DE_TAX_ID:   "DE Tax ID",
    PayloadCategory.FR_INSEE:    "FR INSEE",
    PayloadCategory.AWS_KEY:     "AWS Key",
    PayloadCategory.GITHUB_TOKEN:"GitHub Token",
    PayloadCategory.STRIPE_KEY:  "Stripe Key",
    PayloadCategory.SLACK_TOKEN: "Slack Token",
    PayloadCategory.JWT:         "JWT",
    PayloadCategory.CLASSIFICATION: "Classification",
    PayloadCategory.EMAIL:       "Email",
    PayloadCategory.PHONE:       "Phone",
}


def _sheet_name(cat: PayloadCategory) -> str:
    title = _SECTION_TITLES.get(cat, cat.value)
    return title[:31]  # Excel sheet name limit


def _style_header_row(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 12, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


def _build_summary(ws, by_cat: dict, total: int, today: str) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "DLP Test Document — Evadex Generate"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws["A2"] = "Generated"
    ws["B2"] = today
    ws["A3"] = "Total entries"
    ws["B3"] = total

    evasion_count = sum(
        1 for cat_entries in by_cat.values()
        for e in cat_entries if e.technique is not None
    )
    ws["A4"] = "Evasion variants"
    ws["B4"] = evasion_count
    ws["A5"] = "Plain values"
    ws["B5"] = total - evasion_count

    ws["A7"] = "Category"
    ws["B7"] = "Count"
    ws["C7"] = "Evasions"
    ws["D7"] = "Evasion %"
    _style_header_row(ws, 4)

    for i, (cat, cat_entries) in enumerate(
        sorted(by_cat.items(), key=lambda kv: kv[0].value), start=8
    ):
        evasions = sum(1 for e in cat_entries if e.technique is not None)
        pct = round(evasions / len(cat_entries) * 100, 1) if cat_entries else 0.0
        ws.cell(row=i, column=1).value = _SECTION_TITLES.get(cat, cat.value)
        ws.cell(row=i, column=2).value = len(cat_entries)
        ws.cell(row=i, column=3).value = evasions
        ws.cell(row=i, column=4).value = f"{pct}%"


def write_xlsx(entries: list[GeneratedEntry], path: str) -> None:
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    by_cat: dict[PayloadCategory, list[GeneratedEntry]] = defaultdict(list)
    for e in entries:
        by_cat[e.category].append(e)

    today = datetime.date.today().isoformat()

    # Summary sheet
    _build_summary(default_sheet, by_cat, len(entries), today)

    headers = [
        "#", "Embedded Text", "Plain Value",
        "Variant Value", "Technique", "Generator", "Has Keywords",
    ]

    for cat in sorted(by_cat.keys(), key=lambda c: c.value):
        cat_entries = by_cat[cat]
        ws = wb.create_sheet(_sheet_name(cat))

        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, len(headers))
        ws.freeze_panes = "A2"

        for row_idx, e in enumerate(cat_entries, 2):
            fill = _EVASION_FILL if e.technique else (
                _ALT_FILL if row_idx % 2 == 0 else None
            )
            cells = [
                (1, row_idx - 1),
                (2, e.embedded_text),
                (3, e.plain_value),
                (4, e.variant_value),
                (5, e.technique or ""),
                (6, e.generator_name or ""),
                (7, str(e.has_keywords)),
            ]
            for col, value in cells:
                cell = ws.cell(row=row_idx, column=col, value=value)
                if fill:
                    cell.fill = fill
                if col == 2:
                    cell.alignment = Alignment(wrap_text=True)

        _auto_width(ws)
        ws.row_dimensions[1].height = 18

    wb.save(path)
